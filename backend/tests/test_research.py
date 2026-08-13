from __future__ import annotations

import asyncio
from pathlib import Path
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.config import Settings, get_settings
from app.main import app
from app.services.duimp_policy import prepare_product
from app.services.research.enrichment import EvidenceEnrichmentService
from app.services.research.filtering import canonical_url, evaluate_result, score_result
from app.services.research.schemas import EnrichmentResponse, ResearchResponse
from app.services.research.selection import select_rich_products
from app.services.research.service import ProductResearchService
from app.services.spreadsheets.schemas import Product


def make_product(code: str = "NB-690") -> Product:
    product = Product(
        product_id=code,
        code=code,
        code_original=code,
        code_confidence=0.99,
        sheet_name="Packing",
        row_numbers=[2],
        item_name="Conjunto feminino tecido plano",
        ncm="6104.23.00",
        composition="95% poliéster 5% elastano",
    )
    prepare_product(product)
    return product


def gateway_payload(*results: dict) -> dict:
    return {"provider": "searxng-search", "results": list(results), "errors": []}


def raw_result(title: str, url: str, snippet: str, position: int = 1) -> dict:
    return {
        "title": title,
        "url": url,
        "snippet": snippet,
        "position": position,
        "metadata": {"source_type": "brave"},
        "citation": {
            "provider": "searxng-search",
            "retrieved_at": "2026-08-13T18:29:09.327Z",
        },
    }


def test_filter_rejects_ambiguous_code_noise_and_keeps_corroborated_result() -> None:
    ambiguous = make_product("WW77#")
    noise = raw_result(
        "World at War, Issue #77",
        "https://shop.example.com/game?utm_source=test",
        "WW77 strategy board game",
    )
    assert score_result(noise, ambiguous, "WW77", "searxng-search") is None

    product = make_product()
    relevant = raw_result(
        "NB-690 conjunto feminino",
        "https://supplier.example/catalog/nb-690?utm_source=search&color=blue#photo",
        "Conjunto em tecido plano, composição 95% poliéster e 5% elastano.",
    )
    evidence = score_result(relevant, product, '"NB-690" "Conjunto feminino"', "searxng-search")
    assert evidence is not None
    assert evidence.url == "https://supplier.example/catalog/nb-690?color=blue"
    assert evidence.score >= 7
    assert evidence.evidence_strength == "STRONG"
    assert "código/modelo presente na URL" in evidence.relevance_reasons


def test_filter_blocks_social_and_invalid_urls() -> None:
    product = make_product()
    social = raw_result("NB-690 conjunto feminino", "https://youtube.com/watch?v=1", "95% poliéster")
    assert score_result(social, product, "NB-690", "searxng-search") is None
    assert canonical_url("javascript:alert(1)") is None


def test_filter_rejects_query_reflection_with_unrelated_snippet() -> None:
    product = make_product("CY2926")
    reflected = raw_result(
        'Texto aleatório "CY2926" "Conjunto feminino tecido plano" NCM 6104.23.00',
        "http://random.example/other",
        "Palavras sem relação com o produto ou com sua composição.",
    )
    reflected["metadata"] = {"source_type": "qwant"}
    evaluation = evaluate_result(
        reflected, product, '"CY2926" "Conjunto feminino tecido plano"', "searxng-search"
    )
    assert evaluation.evidence is None
    assert evaluation.discard_reason == "query_echo"


def test_official_looking_domain_is_not_invented_but_can_supply_moderate_evidence() -> None:
    product = make_product("MODEL-900")
    product.manufacturer = "Acme"
    relevant = raw_result(
        "MODEL-900 conjunto feminino",
        "https://acme.example/catalog/item",
        "Conjunto feminino em tecido plano.",
    )

    evidence = score_result(relevant, product, '"MODEL-900" "conjunto feminino"', "searxng-search")

    assert evidence is not None
    assert evidence.source_category == "MANUFACTURER"
    assert evidence.evidence_strength == "MODERATE"
    assert any("oficialidade não confirmada" in reason for reason in evidence.relevance_reasons)


def test_selects_rich_products_while_maximizing_metadata_diversity() -> None:
    products = []
    for code, item, composition, manufacturer in (
        ("LONG-CODE-01", "Women's top", "polyester", "Factory A"),
        ("LONG-CODE-02", "Women's top", "polyester", "Factory A"),
        ("LONG-CODE-03", "Dress", "cotton", "Factory B"),
    ):
        product = make_product(code)
        product.item_name = item
        product.composition = composition
        product.construction = "woven"
        product.manufacturer = manufacturer
        prepare_product(product)
        products.append(product)

    selected = select_rich_products(products, limit=2)

    assert [item.product.product_id for item in selected] == ["LONG-CODE-01", "LONG-CODE-03"]
    assert "manufacturer" in selected[0].available_fields


def test_research_deduplicates_and_uses_persistent_query_cache(tmp_path: Path) -> None:
    product = make_product()

    class FakeGateway:
        calls = 0

        async def search(self, query: str, *, provider: str, max_results: int) -> dict:
            self.calls += 1
            return gateway_payload(
                raw_result("NB-690 conjunto feminino", "https://supplier.example/nb-690?utm_source=a", "95% poliéster 5% elastano"),
                raw_result("NB-690 conjunto feminino", "https://supplier.example/nb-690", "Conjunto em tecido plano", 2),
                raw_result("NB-690 video", "https://youtube.com/watch?v=1", "produto", 3),
            )

    settings = Settings(
        upload_dir=tmp_path,
        search_cache_dir=tmp_path / "cache",
        search_cache_ttl_seconds=3600,
        omniroute_max_retries=0,
    )
    gateway = FakeGateway()
    service = ProductResearchService(settings, gateway=gateway)  # type: ignore[arg-type]
    first = asyncio.run(service.research(
        "file-1", [product], max_queries_per_product=1, max_results_per_query=8, refresh_cache=False
    ))
    second = asyncio.run(service.research(
        "file-1", [product], max_queries_per_product=1, max_results_per_query=8, refresh_cache=False
    ))

    assert gateway.calls == 1
    assert first.gateway_calls == 1
    assert first.products[0].status == "OK"
    assert len(first.products[0].evidences) == 1
    assert first.products[0].raw_results == 3
    assert first.products[0].deduplicated_results == 2
    assert first.products[0].discarded_results == 2
    assert first.products[0].discard_reasons == {"duplicate_url": 1, "blocked_domain": 1}
    assert second.gateway_calls == 0
    assert second.cache_hits == 1
    assert second.cache_misses == 0
    assert second.products[0].queries[0].from_cache is True
    assert second.products[0].queries[0].cache_status == "HIT"
    assert second.products[0].queries[0].provider == "searxng-search"
    assert second.llm_used is False


def test_positive_control_uses_the_regular_pipeline_and_keeps_official_product_evidence(
    tmp_path: Path,
) -> None:
    product = Product(
        product_id="POSITIVE_CONTROL",
        code="Raspberry Pi 5",
        code_original="Raspberry Pi 5",
        code_confidence=1.0,
        sheet_name="POSITIVE_CONTROL",
        row_numbers=[1],
        item_name="single-board computer",
        manufacturer="Raspberry Pi",
        brand="Raspberry Pi",
    )
    prepare_product(product)

    class FakeGateway:
        calls = 0

        async def search(self, query: str, *, provider: str, max_results: int) -> dict:
            self.calls += 1
            result = raw_result(
                "Buy a Raspberry Pi 5 – Raspberry Pi",
                "https://www.raspberrypi.com/products/raspberry-pi-5/?utm_source=test",
                "Raspberry Pi 5 is a single-board computer from Raspberry Pi.",
            )
            result["metadata"] = {"source_type": "bing"}
            return gateway_payload(result)

    settings = Settings(
        upload_dir=tmp_path,
        search_cache_dir=tmp_path / "cache",
        search_cache_ttl_seconds=3600,
        omniroute_max_retries=0,
    )
    gateway = FakeGateway()
    response = asyncio.run(ProductResearchService(settings, gateway=gateway).research(
        "POSITIVE_CONTROL",
        [product],
        max_queries_per_product=3,
        max_results_per_query=8,
        refresh_cache=False,
    ))

    assert product.research_preparation.queries[:3] == [
        '"Raspberry Pi 5"',
        '"Raspberry Pi 5" "single-board computer"',
        '"Raspberry Pi 5" "Raspberry Pi"',
    ]
    assert gateway.calls == 3
    assert response.products[0].status == "OK"
    assert len(response.products[0].evidences) == 1
    evidence = response.products[0].evidences[0]
    assert evidence.url == "https://www.raspberrypi.com/products/raspberry-pi-5/"
    assert evidence.source_category == "MANUFACTURER"
    assert evidence.evidence_strength == "STRONG"
    assert evidence.source_engine == "bing"
    assert response.llm_used is False


def test_research_endpoint_limits_pilot_to_two_or_three_products(tmp_path: Path) -> None:
    file_id = str(uuid4())
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Item name", "NCM"])
    sheet.append(["ONE-1", "Produto", "6104.23.00"])
    workbook.save(tmp_path / f"{file_id}.xlsx")
    workbook.close()
    settings = Settings(upload_dir=tmp_path, search_cache_dir=tmp_path / "cache")
    app.dependency_overrides[get_settings] = lambda: settings
    try:
        response = TestClient(app).post(
            f"/api/uploads/{file_id}/research",
            json={"product_ids": ["ONE-1"]},
        )
    finally:
        app.dependency_overrides.clear()
    assert response.status_code == 422


def test_research_endpoint_runs_two_products_and_keeps_llm_disabled(
    tmp_path: Path, monkeypatch,
) -> None:
    file_id = str(uuid4())
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Item name", "NCM"])
    sheet.append(["ONE-1", "Produto um", "6104.23.00"])
    sheet.append(["TWO-2", "Produto dois", "6104.43.00"])
    workbook.save(tmp_path / f"{file_id}.xlsx")
    workbook.close()
    settings = Settings(upload_dir=tmp_path, search_cache_dir=tmp_path / "cache")
    captured = {}

    async def fake_research(self, requested_file_id, products, **options):
        captured["ids"] = [product.product_id for product in products]
        captured["options"] = options
        return ResearchResponse(
            file_id=requested_file_id,
            provider="searxng-search",
            researched_at="2026-08-13T18:29:09Z",
            products=[],
            query_count=0,
            gateway_calls=0,
            cache_hits=0,
            cache_misses=0,
            llm_used=False,
        )

    monkeypatch.setattr(ProductResearchService, "research", fake_research)
    app.dependency_overrides[get_settings] = lambda: settings
    try:
        response = TestClient(app).post(
            f"/api/uploads/{file_id}/research",
            json={"product_ids": ["ONE-1", "TWO-2"], "max_queries_per_product": 4},
        )
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 200
    assert response.json()["llm_used"] is False
    assert captured["ids"] == ["ONE-1", "TWO-2"]
    assert captured["options"]["max_queries_per_product"] == 4


def test_enrichment_endpoint_preserves_research_contract_and_limits_pages(
    tmp_path: Path, monkeypatch,
) -> None:
    file_id = str(uuid4())
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", "Item name"])
    sheet.append(["ONE-1", "Produto um"])
    sheet.append(["TWO-2", "Produto dois"])
    workbook.save(tmp_path / f"{file_id}.xlsx")
    workbook.close()
    settings = Settings(
        upload_dir=tmp_path,
        search_cache_dir=tmp_path / "search-cache",
        fetch_cache_dir=tmp_path / "fetch-cache",
    )
    captured = {}

    async def fake_research(self, requested_file_id, products, **options):
        captured["research_ids"] = [product.product_id for product in products]
        return ResearchResponse(
            file_id=requested_file_id, provider="searxng-search",
            researched_at="2026-08-13T18:29:09Z", products=[], query_count=0,
            gateway_calls=0, cache_hits=0, cache_misses=0, llm_used=False,
        )

    async def fake_enrich(self, requested_file_id, products, research, **options):
        captured["max_pages"] = options["max_pages_per_product"]
        captured["refresh_fetch_cache"] = options["refresh_fetch_cache"]
        return EnrichmentResponse(
            file_id=requested_file_id, provider="searxng-search",
            researched_at="2026-08-13T18:29:09Z", research=research, products=[],
            fetch_requests=0, fetch_cache_hits=0, fetch_cache_misses=0,
            fetch_cache_expired=0, llm_used=False,
        )

    monkeypatch.setattr(ProductResearchService, "research", fake_research)
    monkeypatch.setattr(EvidenceEnrichmentService, "enrich", fake_enrich)
    app.dependency_overrides[get_settings] = lambda: settings
    try:
        response = TestClient(app).post(
            f"/api/uploads/{file_id}/research/enrich",
            json={
                "product_ids": ["ONE-1", "TWO-2"],
                "max_pages_per_product": 3,
                "refresh_fetch_cache": True,
            },
        )
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 200
    assert response.json()["llm_used"] is False
    assert captured["research_ids"] == ["ONE-1", "TWO-2"]
    assert captured["max_pages"] == 3
    assert captured["refresh_fetch_cache"] is True
