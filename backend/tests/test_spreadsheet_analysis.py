from __future__ import annotations

import json
from pathlib import Path
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from openpyxl.drawing.image import Image as SpreadsheetImage
from PIL import Image

from app.config import Settings, get_settings
from app.main import app
from app.services.spreadsheets import analyze_workbook


client = TestClient(app)


def save_workbook(
    path: Path,
    headers: list[str],
    rows: list[list[object]],
    *,
    header_row: int = 1,
    sheet_name: str = "Packing",
    extra_sheet: bool = False,
    merges: list[str] | None = None,
    images: list[tuple[str, str]] | None = None,
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for column, value in enumerate(headers, 1):
        sheet.cell(header_row, column, value)
    for row_number, values in enumerate(rows, header_row + 1):
        for column, value in enumerate(values, 1):
            sheet.cell(row_number, column, value)
    for reference in merges or []:
        sheet.merge_cells(reference)
    for index, (anchor, kind) in enumerate(images or [], 1):
        color = {"product": "red", "wash": "blue", "hangtag": "green", "label": "yellow", "other": "purple"}[kind]
        image_path = path.parent / f"image-{index}.png"
        Image.new("RGB", (24 + index, 16 + index), color=color).save(image_path)
        picture = SpreadsheetImage(image_path)
        picture.anchor = anchor
        sheet.add_image(picture)
    if extra_sheet:
        workbook.create_sheet("Notas")["A1"] = "sem tabela"
    workbook.save(path)
    workbook.close()
    return path


@pytest.mark.parametrize("header", ["Style number", "Code", "款号"])
def test_detects_known_code_headers(tmp_path: Path, header: str) -> None:
    path = save_workbook(
        tmp_path / f"known-{uuid4()}.xlsx",
        ["Picture", header, "Item name", "NCM"],
        [[None, "ABC-123", "Produto", "6104.43.00"]],
    )
    result = analyze_workbook(path, str(uuid4()))
    assert result.code_column == 2
    assert result.code_confidence == 0.99
    assert result.products[0].code == "ABC-123"


def test_detects_header_away_from_first_row_and_multiple_sheets(tmp_path: Path) -> None:
    path = save_workbook(
        tmp_path / "later-header.xlsx",
        ["Picture", "Product Code", "Composition"],
        [[None, "CY2927", "100% poliéster"]],
        header_row=4,
        sheet_name="Produtos",
        extra_sheet=True,
    )
    result = analyze_workbook(path, str(uuid4()))
    assert result.main_sheet == "Produtos"
    assert result.sheets == 2
    assert result.header_rows == [4]
    assert result.products[0].composition == "100% poliéster"


def test_detects_multiline_header(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Multiline"
    sheet["A1"] = "Picture"
    sheet["B2"] = "Style number"
    sheet["C2"] = "Ingredients"
    sheet["B3"] = "ML-100"
    sheet["C3"] = "100% algodão"
    path = tmp_path / "multiline.xlsx"
    workbook.save(path)
    workbook.close()

    result = analyze_workbook(path, str(uuid4()))
    assert result.header_rows == [1, 2]
    assert result.code_column == 2
    assert result.products[0].code == "ML-100"


def test_infers_code_column_from_values_when_title_is_unknown(tmp_path: Path) -> None:
    path = save_workbook(
        tmp_path / "inferred.xlsx",
        ["Item name", "Identificador interno", "NCM"],
        [["Produto A", "CY2927", "6104.43.00"], ["Produto B", "CY2926", "6104.42.00"]],
    )
    result = analyze_workbook(path, str(uuid4()))
    assert result.code_column == 2
    assert 0.5 <= result.code_confidence < 0.99
    assert {product.code for product in result.products} == {"CY2927", "CY2926"}


def test_missing_code_creates_review_identifier_without_inventing(tmp_path: Path) -> None:
    path = save_workbook(
        tmp_path / "missing-code.xlsx",
        ["Picture", "Item name", "NCM"],
        [[None, "Produto sem código", "6104.43.00"]],
    )
    result = analyze_workbook(path, str(uuid4()))
    product = result.products[0]
    assert product.product_id.startswith("ROW-")
    assert product.code is None
    assert product.status == "REVISAR"
    assert "Código não identificado" in product.warnings


def test_normalizes_logistics_and_groups_repeated_codes(tmp_path: Path) -> None:
    path = save_workbook(
        tmp_path / "repeated.xlsx",
        ["图片", "款号", "品名", "箱数"],
        [[None, "WW77# （1-21#箱）", "连衣裙", 21], [None, "WW77# （22#箱）", "连衣裙", 1]],
    )
    result = analyze_workbook(path, str(uuid4()))
    assert result.unique_products == 1
    product = result.products[0]
    assert product.code == "WW77#"
    assert product.code_original == "WW77# （1-21#箱）"
    assert product.row_numbers == [2, 3]
    assert product.original_values["code"] == ["WW77# （1-21#箱）", "WW77# （22#箱）"]
    assert result.repeated_codes == {"WW77#": [2, 3]}


def test_preserves_non_logistical_chinese_suffix_in_code(tmp_path: Path) -> None:
    path = save_workbook(
        tmp_path / "semantic-suffix.xlsx",
        ["款号", "品名", "NCM"],
        [["TP173#D", "梭织中裤", "6104.63.00"], ["TP173#D腰带", "腰带", "3926.20.00"]],
    )
    result = analyze_workbook(path, str(uuid4()))
    assert {product.code for product in result.products} == {"TP173#D", "TP173#D腰带"}


def test_merged_code_applies_to_all_logical_rows(tmp_path: Path) -> None:
    path = save_workbook(
        tmp_path / "merged.xlsx",
        ["Picture", "Style", "Item name", "NCM"],
        [[None, "MERGED-1", "Produto", "6104.43.00"], [None, None, "Produto", "6104.43.00"]],
        merges=["B2:B3"],
    )
    result = analyze_workbook(path, str(uuid4()))
    assert result.unique_products == 1
    assert result.products[0].row_numbers == [2, 3]


def test_extracts_classifies_and_associates_images_without_base64(tmp_path: Path) -> None:
    path = save_workbook(
        tmp_path / "images.xlsx",
        ["Picture", "Style", "Item name", "Wash Label", "Hangtag", "Label", "Auxiliar"],
        [
            [None, "IMG-1", "Produto", None, None, None, None],
            [None, None, "Sem código", None, None, None, None],
        ],
        images=[("A2", "product"), ("D2", "wash"), ("E2", "hangtag"), ("F2", "label"), ("G3", "other")],
    )
    result = analyze_workbook(path, str(uuid4()))
    assert result.images_detected == 5
    assert result.product_images == 1
    assert result.wash_labels == 1
    assert result.hangtags == 1
    assert result.label_images == 1
    assert result.other_images == 1
    coded = next(product for product in result.products if product.code == "IMG-1")
    assert coded.images.product[0].related_code == "IMG-1"
    assert coded.images.product[0].width == 25
    review = next(product for product in result.products if product.code is None)
    assert review.status == "REVISAR"
    assert len(review.images.other) == 1
    payload = json.dumps(result.model_dump(mode="json"))
    assert "base64" not in payload.casefold()
    assert "data:image" not in payload.casefold()


def test_endpoint_validates_uuid_and_never_accepts_arbitrary_path(tmp_path: Path) -> None:
    file_id = uuid4()
    save_workbook(
        tmp_path / f"{file_id}.xlsx",
        ["Picture", "Style number", "Item name"],
        [[None, "API-1", "Produto"]],
    )
    settings = Settings(upload_dir=tmp_path)
    app.dependency_overrides[get_settings] = lambda: settings
    try:
        response = client.post(f"/api/uploads/{file_id}/analyze")
        assert response.status_code == 200
        assert response.json()["file_id"] == str(file_id)
        serialized = response.text
        assert str(tmp_path) not in serialized
        assert "base64" not in serialized.casefold()
        assert client.post("/api/uploads/not-a-uuid/analyze").status_code == 422
        assert client.post(f"/api/uploads/{uuid4()}/analyze").status_code == 404
        assert client.post("/api/uploads/../../etc/passwd/analyze").status_code in {404, 422}
    finally:
        app.dependency_overrides.clear()
